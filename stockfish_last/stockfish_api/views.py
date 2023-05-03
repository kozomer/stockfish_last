from django.shortcuts import render
import pandas as pd
from .models import (Customers, Products, Sales, Warehouse, ROP, Salers, SalerPerformance, SaleSummary, SalerMonthlySaleRating, 
                    MonthlyProductSales,CustomerPerformance, ProductPerformance, OrderList, GoodsOnRoad, Trucks, NotificationsOrderList)
from django.views import View
from rest_framework.views import APIView
from django.http import JsonResponse, HttpResponse
import json
from django.db.models.signals import post_save, post_delete, pre_save
from django.dispatch import receiver
from .definitions import jalali_to_greg, greg_to_jalali, calculate_experience_rating, calculate_sale_rating, calculate_passive_saler_experience_rating, current_jalali_date, get_exchange_rate, get_model, generate_future_forecast_dates, the_man_from_future
from datetime import datetime
import datetime
import jdatetime
from django.db.models import Sum, Avg
from django.db.models.functions import Coalesce
from django.views.decorators.csrf import csrf_exempt
import statistics
import math
from scipy.stats import norm
import traceback
from django.db import models
import openpyxl
import base64
from io import BytesIO
import numpy as np
from django.db import transaction



# Authentications
from django.contrib.auth import authenticate
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.decorators import  permission_classes, authentication_classes
from django.db.utils import OperationalError
from rest_framework.exceptions import ValidationError
import filetype




# region Login/Logout

class LoginView(TokenObtainPairView):
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        username = data.get('username')
        username = username.strip()
        password = data.get('password')
       
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            response = super().post(request, *args, **kwargs)
            # Convert the first name and last name to uppercase
            first_name = user.first_name.upper()
            last_name = user.last_name.upper()
            # Add the uppercase names to the response
            response.data['first_name'] = first_name
            response.data['last_name'] = last_name
            return response
        else:
            return JsonResponse({'error': 'Invalid credentials'}, status=401)

class LogoutView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request):
        try:
            refresh_token = request.POST.get('refresh_token')
            token = RefreshToken(refresh_token)
            token.blacklist()

            return JsonResponse({'success': 'successfully log out'}, status=205)
        except Exception as e:
            return JsonResponse({'error': 'BAD REQUEST'}, status=400)


# endregion

# region Customers

class AddCustomersView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    
    def post(self, request, *args, **kwargs):
        try:
            if 'file' not in request.FILES:
                raise ValidationError("No file uploaded")
            file = request.FILES['file']
            kind = filetype.guess(file.read())
            if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                return JsonResponse({'error': "The uploaded file is not a valid Excel file!"}, status=400)

            data = pd.read_excel(file)
            if data.empty:
                return JsonResponse({'error': "The uploaded file is empty"}, status=400)
            count = 0
            for i, row in data.iterrows():
                try:
                    customer_code = row["Customer Code"]
                    if Customers.objects.filter(customer_code=customer_code).exists():
                        continue
                    count+=1
                    customer = Customers(customer_code=customer_code, description=row["Description"], quantity=row["Quantity"],
                                        area_code=row["Area Code"], code=row["Code"], city=row["City"], area=row["Area"])
                    customer.save()
                except KeyError as e:
                    return JsonResponse({'error': f"Column '{e}' not found in the uploaded file"}, status=400)
            return JsonResponse({'message': f"{count} Customers data added successfully"}, status=200)
        except OperationalError as e:
            return JsonResponse({'error': f"Database error: {str(e)}"}, status=500)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class ViewCustomersView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def get(self,request,*args, **kwargs):
         customers = Customers.objects.values().all()
         customer_list = [[customer['customer_code'], customer['description'], customer['quantity'],
                      customer['area_code'], customer['code'], customer['city'], customer['area']]
                     for customer in customers]
         return JsonResponse(customer_list,safe=False)

class DeleteCustomerView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        try:
            customer_code = request.POST.get('customer_code')
            Customers.objects.filter(customer_code=customer_code).delete()
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        return JsonResponse({'message': "Customer object has been successfully deleted"}, status=200)


class EditCustomerView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            # Check if old_customer_code is provided
            old_customer_code = data.get('old_customer_code')
            if not old_customer_code:
                return JsonResponse({'error': "Missing required parameter: old_customer_code"}, status=400)

            # Get the customer object
            customer = Customers.objects.get(customer_code=old_customer_code)

            # Check if new_customer_code value is unique and not empty
            new_customer_code = data.get('new_customer_code')
            if new_customer_code and new_customer_code != old_customer_code:
                if not new_customer_code:
                    return JsonResponse({'error': "Customer Code cannot be empty!"}, status=400)
                if Customers.objects.filter(customer_code=new_customer_code).exists():
                    return JsonResponse({'error': f"The customer code '{new_customer_code}' already exists in the database."}, status=400)
                else:
                    customer.customer_code = new_customer_code

            # Update other customer fields
            try:
                for field in ['description', 'quantity', 'area_code', 'code', 'city', 'area']:
                    value = data.get(f'new_{field}')
                    if value is not None and value != '':
                        setattr(customer, field, value)
                    else: 
                        return JsonResponse({'error': "One or more data field is empty!"}, status=400)
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=400)


            customer.save()
            return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)

        except Customers.DoesNotExist:
            return JsonResponse({'error': "Customer not found!"}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class ExportCustomersView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        def set_column_widths(worksheet):
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        customers = Customers.objects.all().values()
        # Create a new workbook and add a worksheet
        jalali_date= current_jalali_date().strftime('%Y-%m-%d')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Customers {jalali_date}"
        # Write the header row
        header = ['Customer Code', 'Description', 'Quantity', 'Area Code', 'Code', 'City', 'Area']
        for col_num, column_title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
            cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'),
                                                 bottom=openpyxl.styles.Side(style='medium'),
                                                 left=openpyxl.styles.Side(style='medium'),
                                                 right=openpyxl.styles.Side(style='medium'))
        # Write the data rows
        for row_num, customer in enumerate(customers, 2):
            row = [customer['customer_code'], customer['description'], customer['quantity'],
                   customer['area_code'], customer['code'], customer['city'], customer['area']]
            for col_num, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
        # Apply some styling to the Excel file
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'),
                                                     left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'))
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        # Set the column widths
        set_column_widths(ws)

        # Apply auto filter
        ws.auto_filter.ref = f"A1:G{ws.max_row}"


        # Set the response headers for an Excel file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.read()

        # Encode the content in base64
        base64_content = base64.b64encode(content).decode()


        # Send the content and filename in the JSON response
        return JsonResponse({'filename': f'customers({jalali_date}).xlsx', 'content': base64_content})


        
    

# endregion

# region Sales

class AddSalesView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({'error': "No file uploaded"}, status=400)
            
            file = request.FILES['file']
            kind = filetype.guess(file.read())
            
            if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                return JsonResponse({'error': "The uploaded file is not a valid Excel file"}, status=400)

            data = pd.read_excel(file)
            if data.empty:
                return JsonResponse({'error': "The uploaded file is empty"}, status=400)

            count = 0
            for i, row in data.iterrows():
                no = row["No"]
                if Sales.objects.filter(no=no).exists():
                    continue
                count += 1

                # Check required fields
                for field in ['Good Code', 'Customer Code', 'Original Output Value', 'Net Sales', 'Saler', 'PSR']:
                    if not row[field]:
                        return JsonResponse({'error': f"{field} cannot be empty"}, status=400)

                # Check integer fields
                for field in ['Good Code', 'Customer Code']:
                    try:
                        if not isinstance(int(row[field]), int):
                            return JsonResponse({'error': f"{field} should be integer"}, status=400)
                    except Exception:
                            return JsonResponse({'error': f"{field} should be integer"}, status=400)


                # Check valid date format
                try:
                    date = jdatetime.date(int(row["Year"]), int(row["Month"]), int(row["Day"]))
                except ValueError:
                    return JsonResponse({'error': "Date should be in the format of YYYY-MM-DD"}, status=400)
                except IndexError as e:
                    return JsonResponse({'error': "Date should be in the format of YYYY-MM-DD"}, status=400)
                except Exception as e:
                    return JsonResponse({'error': "Date should be in the format of YYYY-MM-DD"}, status=400)

                # Check valid psr value
                if row['PSR'] not in ['P', 'S', 'R']:
                    return JsonResponse({'error': "Invalid P-S-R value. Allowed values are 'P', 'S', and 'R'."}, status=400)

                # Check for existing good
                if not Warehouse.objects.filter(product_code=row['Good Code']).exists():
                    return JsonResponse({'error': f"No product found with code '{row['Good Code']}' in warehouse"}, status=400)
                    

                # Check for existing customer
                if not Customers.objects.filter(customer_code=row['Customer Code']).exists():
                    return JsonResponse({'error': f"No customer found with code '{row['Customer Code']}'"}, status=400)

                # Check for existing saler
                if not Salers.objects.filter(name=row['Saler']).exists():
                    return JsonResponse({'error': f"No saler found with name '{row['Saler']}'"}, status=400)

                try:
                    customer = Customers.objects.get(customer_code= row["Customer Code"] )
                except Exception as e:
                    return JsonResponse({'error': "No customer found"}, status=400)
                try:
                    product = Products.objects.get(product_code_ir= row["Good Code"] )
                except Exception as e:
                    return JsonResponse({'error': "No product found"}, status=400)
                try:
                    saler = Salers.objects.get(name= row["Saler"] )
                except Exception as e:
                    
                    return JsonResponse({'error': "No saler found"}, status=400)
                
                # Save the Sale object
                sale = Sales(
                    no=no,
                    bill_number=row["Bill Number"],
                    date=date,
                    psr=row["PSR"],
                    customer_code=row["Customer Code"],
                    name= customer.description,
                    city= customer.city,
                    area= customer.area,
                    color_making_saler=row["Color Making Saler"],
                    group= product.group,
                    product_code=row["Good Code"],
                    product_name=product.description_ir,
                    unit=product.unit,
                    unit2=product.unit_secondary,
                    original_value=row["The Original Value"],
                    kg = row['KG'],
                    original_output_value=row["Original Output Value"],
                    secondary_output_value=row["Secondary Output Value"],
                    price=row["Price"],
                    original_price=row["Original Price"],
                    discount_percentage=row["Discount Percantage (%)"],
                    amount_sale=row["Amount Sale"],
                    discount=row["Discount"],
                    additional_sales=row["Additional Sales"],
                    net_sales=row["Net Sales"],
                    discount_percentage_2=row["Discount Percantage 2(%)"],
                    real_discount_percentage=row["Real Discount Percantage (%)"],
                    payment_cash=row["Payment Cash"],
                    payment_check=row["Payment Check"],
                    balance=row["Balance"],
                    saler=row["Saler"],
                    currency_sepidar=row["Currency-Sepidar"],
                    dollar_sepidar=row["Dollar-Sepidar"],
                    currency=row["Currency"],
                    dollar=row["Dollar"],
                    manager_rating=row["Manager Rating"],
                    senior_saler=row["Senior Saler"],
                    tot_monthly_sales=row["Tot Monthly Sales"],
                    receipment=row["Receipment"],
                    ct=row["CT"],
                    payment_type=row["Payment Type"],
                    customer_size=row["Customer Size"],
                    saler_factor=row["Saler Factor"],
                    prim_percentage=row["Prim Percantage"],
                    bonus_factor=row["Bonus Factor"],
                    bonus=row["Bonus"]
                )
                sale.save()

                # Update stock in warehouse
                try:
                    warehouse_item = Warehouse.objects.get(product_code=sale.product_code)
                    warehouse_item.stock -= sale.original_output_value
                    warehouse_item.save()
                except Warehouse.DoesNotExist:
                    return JsonResponse({'error': f"No product found with code '{row['Good Code']}' in warehouse"}, status=400)

                count += 1
            return JsonResponse({'message': f"{count} sales data added successfully"}, status=200)

        except OperationalError as e:
            return JsonResponse({'error': f"Database error: {str(e)}"}, status=500)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)


class ViewSalesView(APIView):
    permission_classes = [IsAuthenticated,]
    authentication_classes = [JWTAuthentication,]
    def get(self, request, *args, **kwargs):
        sales = Sales.objects.values().all()
        sale_list = [[sale['no'], sale['bill_number'], sale['date'].strftime('%Y-%m-%d'), sale['psr'], sale['customer_code'],
                        sale['name'], sale['city'], sale['area'], sale['color_making_saler'], sale['group'], sale['product_code'], 
                        sale['product_name'], sale['unit'], sale['unit2'], sale['kg'], sale['original_value'], sale['original_output_value'], 
                        sale['secondary_output_value'], sale['price'], sale['original_price'], sale['discount_percentage'], sale['amount_sale'],
                        sale['discount'], sale['additional_sales'], sale['net_sales'], sale['discount_percentage_2'], sale['real_discount_percentage'],
                        sale['payment_cash'], sale['payment_check'], sale['balance'], sale['saler'], sale['currency_sepidar'], sale['dollar_sepidar'], 
                        sale['currency'], sale['dollar'], sale['manager_rating'], sale['senior_saler'], sale['tot_monthly_sales'], sale['receipment'], 
                        sale['ct'], sale['payment_type'], sale['customer_size'], sale['saler_factor'], sale['prim_percentage'], sale['bonus_factor'], 
                        sale['bonus']]
                     for sale in sales]
        return JsonResponse(sale_list, safe=False)

class DeleteSaleView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        try:
            no = request.POST.get('no', None)
            product_code = request.POST.get('product_code', None)
            original_output_value = request.POST.get('original_output_value', None)
            Sales.objects.filter(no=no).delete()
            try:
                warehouse_item = Warehouse.objects.get(product_code=product_code)
                warehouse_item.stock += float(original_output_value)
                warehouse_item.save()
            except Warehouse.DoesNotExist:
                return JsonResponse({'error': f"No product found with code '{product_code}' in warehouse"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
        return JsonResponse({'message': "Sale object has been successfully deleted"}, status=200)

class EditSaleView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            # Check for required fields
            for field in ['new_product_code', 'new_customer_code', 'new_original_output_value', 'new_net_sales', 'new_saler', 'new_psr', 'new_date']:
                if not data.get(field):
                    return JsonResponse({'error': f"{field} cannot be empty"}, status=400)

            # Check for integer fields
            for field in ['new_product_code', 'new_customer_code']:
                try:
                    if not isinstance(int(data.get(field)), int):
                        return JsonResponse({'error': f"{field} should be integer"}, status=400)
                except Exception as e:
                    return JsonResponse({'error': f"{field} should be integer"}, status=400)

            # Check for valid date format
            try:
                new_date = data.get('new_date').split("-")
                date = jdatetime.date(int(new_date[0]), int(new_date[1]), int(new_date[2]))
            except ValueError:
                return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
            except IndexError as e:
                return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=400)
            # Check for valid psr value
            if data.get('new_psr') not in ['P', 'S', 'R']:
                return JsonResponse({'error': "Invalid P-S-R value. Allowed values are 'P', 'S', and 'R'."}, status=400)

            # Check for existing good
            if not Warehouse.objects.filter(product_code=data.get('new_product_code')).exists():
                return JsonResponse({'error': f"No product found with code '{data.get('new_product_code')}' in Warehouse. Please check product code. If there is a new product please add firstly to Warehouse."}, status=400)
            if not Products.objects.filter(product_code_ir=data.get('new_product_code')).exists():
                return JsonResponse({'error': f"No product found with code '{data.get('new_product_code')}'. Please check product code. If there is a new product please add firstly to Products."}, status=400)
            # Check for existing customer
            if not Customers.objects.filter(customer_code=data.get('new_customer_code')).exists():
                return JsonResponse({'error': f"No customer found with code '{data.get('new_customer_code')}'. Please check customer code. If there is a new customer please add firstly to Customers."}, status=400)

            # Check for existing saler
            if not Salers.objects.filter(name=data.get('new_saler')).exists():
                return JsonResponse({'error': f"No saler found with name '{data.get('new_saler')}'. Please check saler name. If there is a new saler please add firstly to Salers."}, status=400)
            
            # Update Sale object
            old_no = data.get('old_no')
            if data.get('new_no') and data.get('new_no') != old_no:
                if Sales.objects.filter(no=data.get('new_no')).exists():
                    error_message = f"The sale no '{data.get('new_no')}' already exists in the database."
                    return JsonResponse({'error': error_message}, status=400)
            sale = Sales.objects.get(no=old_no)
            sale.no = data.get('new_no')
            sale.bill_number = data.get('new_bill_number')
            sale.date = date
            sale.psr = data.get('new_psr')
            sale.customer_code = data.get('new_customer_code')
            sale.name = data.get('new_name')
            sale.area = data.get('new_area')
            sale.city = data.get('new_city')
            sale.color_making_saler = data.get('new_color_making_saler')
            sale.group = data.get('new_group')
            sale.product_code = data.get('new_product_code')
            sale.product_name = data.get('new_product_name')
            sale.unit = data.get('new_unit')
            sale.unit2 = data.get('new_unit2')
            sale.kg = data.get('new_kg')
            sale.original_value = data.get('new_original_value')
            sale.original_output_value = data.get('new_original_output_value')
            sale.secondary_output_value = data.get('new_secondary_output_value')
            sale.price = data.get('new_price')
            sale.original_price = data.get('new_original_price')
            sale.discount_percentage = data.get('new_discount_percentage')
            sale.amount_sale = data.get('new_amount_sale')
            sale.discount = data.get('new_discount')
            sale.additional_sales = data.get('new_additional_sales')
            sale.net_sales = data.get('new_net_sales')
            sale.discount_percentage_2 = data.get('new_discount_percentage_2')
            sale.real_discount_percentage = data.get('new_real_discount_percentage')
            sale.payment_cash = data.get('new_payment_cash')
            sale.payment_check = data.get('new_payment_check')
            sale.balance = data.get('new_balance')
            sale.saler = data.get('new_saler')
            sale.currency_sepidar = data.get('new_currency_sepidar')
            sale.dollar_sepidar = data.get('new_dollar_sepidar')
            sale.currency = data.get('new_currency')
            sale.dollar = data.get('new_dollar')
            sale.manager_rating = data.get('new_manager_rating')
            sale.senior_saler = data.get('new_senior_saler')
            sale.tot_monthly_sales = data.get('new_tot_monthly_sales')
            sale.receipment = data.get('new_receipment')
            sale.ct = data.get('new_ct')
            sale.payment_type = data.get('new_payment_type')
            sale.customer_size = data.get('new_customer_size')
            sale.saler_factor = data.get('new_saler_factor')
            sale.prim_percentage = data.get('new_prim_percentage')
            sale.bonus_factor = data.get('new_bonus_factor')
            sale.bonus = data.get('new_bonus')

            sale.save()

            return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)

        except Sales.DoesNotExist:
            return JsonResponse({'error': "Sale not found!"}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class ExportSalesView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        def set_column_widths(worksheet):
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        sales = Sales.objects.all().values()
        # Create a new workbook and add a worksheet
        jalali_date= current_jalali_date().strftime('%Y-%m-%d')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Sales {jalali_date}"
        # Write the header row
        header = ['No', 'Bill Number', 'Date', 'PSR', 'Customer Code', 'Name', 'City', 'Area', 'Color Making Saler', 'Group',
                  'Product Code', 'Product Name', 'Unit', 'Unit2', 'Kg', 'Original Value', 'Original Output Value', 'Secondary Output Value',
                  'Price', 'Original Price', 'Discount Percentage', 'Amount Sale', 'Discount', 'Additional Sales', 'Net Sales',
                  'Discount Percentage 2', 'Real Discount Percentage', 'Payment Cash', 'Payment Check', 'Balance', 'Saler', 'Currency Sepidar',
                  'Dollar Sepidar', 'Currency', 'Dollar', 'Manager Rating', 'Senior Saler', 'Total Monthly Sales', 'Receipment', 'CT',
                  'Payment Type', 'Customer Size', 'Saler Factor', 'Prim Percentage', 'Bonus Factor', 'Bonus']
        for col_num, column_title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
            cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'),
                                                 bottom=openpyxl.styles.Side(style='medium'),
                                                 left=openpyxl.styles.Side(style='medium'),
                                                 right=openpyxl.styles.Side(style='medium'))
        ## Write the data rows
        for row_num, sale in enumerate(sales, 2):
            row = [sale['no'], sale['bill_number'], sale['date'].strftime('%Y-%m-%d'), sale['psr'], sale['customer_code'],
                   sale['name'], sale['city'], sale['area'], sale['color_making_saler'], sale['group'], sale['product_code'], 
                   sale['product_name'], sale['unit'], sale['unit2'], sale['kg'], sale['original_value'], sale['original_output_value'], 
                   sale['secondary_output_value'], sale['price'], sale['original_price'], sale['discount_percentage'], sale['amount_sale'],
                   sale['discount'], sale['additional_sales'], sale['net_sales'], sale['discount_percentage_2'], sale['real_discount_percentage'],
                   sale['payment_cash'], sale['payment_check'], sale['balance'], sale['saler'], sale['currency_sepidar'], sale['dollar_sepidar'], 
                   sale['currency'], sale['dollar'], sale['manager_rating'], sale['senior_saler'], sale['tot_monthly_sales'], sale['receipment'], 
                   sale['ct'], sale['payment_type'], sale['customer_size'], sale['saler_factor'], sale['prim_percentage'], sale['bonus_factor'], 
                   sale['bonus']]
            for col_num, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
        # Apply some styling to the Excel file
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'),
                                                     left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'))
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        # Set the column widths
        set_column_widths(ws)

        # Apply auto filter
        ws.auto_filter.ref = f"A1:AT{ws.max_row}"


        # Set the response headers for an Excel file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.read()

        # Encode the content in base64
        base64_content = base64.b64encode(content).decode()


        # Send the content and filename in the JSON response
        return JsonResponse({'filename': f'sales({jalali_date}).xlsx', 'content': base64_content})




# endregion

# region Warehouse

class AddWarehouseView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    
    def post(self, request, *args, **kwargs):
        try:
            if 'file' not in request.FILES:
                raise ValidationError("No file uploaded")
            file = request.FILES['file']
            kind = filetype.guess(file.read())
            if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                return JsonResponse({'error': "The uploaded file is not a valid Excel file"}, status=400)

            data = pd.read_excel(file)
            if data.empty:
                return JsonResponse({'error': "The uploaded file is empty"}, status=400)
            count = 0
            for i, row in data.iterrows():
                try:
                    product_code = int(row["Product Code"])
                    try:
                        product = Products.objects.get(product_code_ir = product_code)
                    except Products.DoesNotExist:
                        return JsonResponse({'error': f"Product Code: '{product_code}' not found in the Products Table. If there is no mistake, please add '{product_code}' to the Products Table first. "}, status=400)

                    if Warehouse.objects.filter(product_code=product_code).exists():
                        continue
                    count+=1
                    warehouse_item = Warehouse(product_code=product_code, title=row["Product Title"], unit=row["Unit"], stock=row["Stock"])
                    warehouse_item.save()
                except KeyError as e:
                    return JsonResponse({'error': f"Column '{e}' not found in the uploaded file"}, status=400)
            return JsonResponse({'message': f"{count} Warehouse items added successfully"}, status=200)
        except OperationalError as e:
            return JsonResponse({'error': f"Database error: {str(e)}"}, status=500)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class ViewWarehouseView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def get(self, request, *args, **kwargs):
        # if not request.user.is_authenticated:
        #     return HttpResponse(status=401)
        warehouse_items = Warehouse.objects.values().all()
        warehouse_list = [[item['product_code'], item['title'], item['unit'], item['stock']] for item in warehouse_items]
        return JsonResponse(warehouse_list, safe=False)

class DeleteWarehouseView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        try:
            product_code = request.POST.get('product_code')
            Warehouse.objects.filter(product_code=product_code).delete()
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        return JsonResponse({'message': "Warehouse object has been successfully deleted"}, status=200)


class EditWarehouseView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            # Check if old_product_code is provided
            old_product_code = data.get('old_product_code')
            if not old_product_code:
                return JsonResponse({'error': "Missing required parameter: Product Code"}, status=400)

            # Get the warehouse item object
            warehouse_item = Warehouse.objects.get(product_code=old_product_code)

            # Check if new_product_code value is unique and not empty
            new_product_code = data.get('new_product_code')
            if new_product_code and new_product_code != old_product_code:
                if not new_product_code:
                    return JsonResponse({'error': "Product Code cannot be empty!"}, status=400)
                if Warehouse.objects.filter(product_code=new_product_code).exists():
                    return JsonResponse({'error': f"The product code '{new_product_code}' already exists in the warehouse."}, status=400)
                else:
                    try:
                        product = Products.objects.get(product_code_ir = new_product_code)
                    except Products.DoesNotExist:
                        return JsonResponse({'error': f"Your new Product Code: '{new_product_code}' not found in the Products Table. If there is no mistake, please add '{new_product_code}' to the Products Table first. "}, status=400)
                    warehouse_item.product_code = new_product_code

            # Update other warehouse item fields
            for field in ['new_title', 'new_unit', 'new_stock']:
                value = data.get(field)
                if value is not None and value != '':
                    updated_field = field[4:]  # Remove the "new_" prefix
                    setattr(warehouse_item, updated_field, value)
                else: 
                    return JsonResponse({'error': f"{field} cannot be empty!"}, status=400)


            warehouse_item.save()
            return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)

        except Warehouse.DoesNotExist:
            return JsonResponse({'error': "Warehouse item not found!"}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class ExportWarehouseView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        def set_column_widths(worksheet):
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        warehouse_items = Warehouse.objects.all().values()
        # Create a new workbook and add a worksheet
        jalali_date = current_jalali_date().strftime('%Y-%m-%d')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Warehouse {jalali_date}"
        # Write the header row
        header = ['Product Code', 'Title', 'Unit', 'Stock']
        for col_num, column_title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
            cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'),
                                                 bottom=openpyxl.styles.Side(style='medium'),
                                                 left=openpyxl.styles.Side(style='medium'),
                                                 right=openpyxl.styles.Side(style='medium'))
        # Write the data rows
        for row_num, item in enumerate(warehouse_items, 2):
            row = [item['product_code'], item['title'], item['unit'], item['stock']]
            for col_num, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
        # Apply some styling to the Excel file
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'),
                                                     left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'))
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        # Set the column widths
        set_column_widths(ws)

        # Apply auto filter
        ws.auto_filter.ref = f"A1:D{ws.max_row}"

        # Set the response headers for an Excel file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.read()

        # Encode the content in base64
        base64_content = base64.b64encode(content).decode()

        # Send the content and filename in the JSON response
        return JsonResponse({'filename': f'warehouse({jalali_date}).xlsx', 'content': base64_content})



# endregion

# region Products

class AddProductsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            if 'file' not in request.FILES:
                return JsonResponse({'error': "No file uploaded"}, status=400)
            file = request.FILES['file']
            kind = filetype.guess(file.read())
            if kind is None or kind.mime not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                return JsonResponse({'error': "The uploaded file is not a valid Excel file"}, status=400)

            data = pd.read_excel(file)
            if data.empty:
                return JsonResponse({'error': "The uploaded file is empty"}, status=400)
            count = 0
            for i, row in data.iterrows():
                try:
                    product_code_ir = row["Product Number IR"]
                    if Products.objects.filter(product_code_ir=product_code_ir).exists():
                        continue
                    count+=1
                    product = Products(
                        group=row["Group"],
                        subgroup=row["Subgroup"],
                        feature=row["Feature"],
                        product_code_ir=product_code_ir,
                        product_code_tr=row["Product Number TR"],
                        description_tr=row["Description TR"],
                        description_ir=row["Description IR"],
                        unit=row["Unit"],
                        unit_secondary=row["Unit Secondary"],
                        weight = row["Weight"],
                        currency = row["Currency"],
                        price= row["Price"]
                    )
                    product.save()
                except KeyError as e:
                    return JsonResponse({'error': f"Column '{e}' not found in the uploaded file"}, status=400)
                except Exception as e:
                    return JsonResponse({'error': str(e)}, status=400)
            return JsonResponse({'message': f"{count} Products data added successfully"}, status=200)
        except OperationalError as e:
            return JsonResponse({'error': f"Database error: {str(e)}"}, status=500)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class ViewProductsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def get(self, request, *args, **kwargs):
        products = Products.objects.values().all()
        product_list = [[p['group'], p['subgroup'], p['feature'], p['product_code_ir'], p['product_code_tr'],
                         p['description_tr'], p['description_ir'], p['unit'], p['unit_secondary'],p['weight'],p['currency'], p['price']] for p in products]
        return JsonResponse(product_list, safe=False)

class DeleteProductView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        try:
            product_code_ir = request.POST.get('product_code_ir')
            Products.objects.filter(product_code_ir=product_code_ir).delete()
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        return JsonResponse({'message': "Product object has been successfully deleted"}, status=200)

class EditProductView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            old_product_code_ir = data.get('old_product_code_ir')
            product = Products.objects.get(product_code_ir=old_product_code_ir)

            # Check if new product_code_ir value is unique
            new_product_code_ir = data.get('new_product_code_ir')
            if new_product_code_ir and new_product_code_ir != old_product_code_ir:
                if Products.objects.filter(product_code_ir=new_product_code_ir).exists():
                    return JsonResponse({'error': f"The Product Code IR '{new_product_code_ir}' already exists in the database."}, status=400)
                if not new_product_code_ir:
                    return JsonResponse({'error': "Product Code IR cannot be empty!"}, status=400)
                else:
                    product.product_code_ir = new_product_code_ir

            # Update other product fields
            for field in [ 'new_currency', 'new_description_ir', 'new_description_tr', 'new_feature', 'new_group', 'new_price', 'new_product_code_tr', 'new_subgroup', 'new_unit', 'new_unit_secondary', 'new_weight']:
                value = data.get(field)
                if value is not None and value != '':
                    updated_field = field[4:]  # Remove the "new_" prefix
                    setattr(product, updated_field, value)
                else:
                    return JsonResponse({'error': f"The field '{field}' cannot be empty."}, status=400)

            product.save()
            return JsonResponse({'message': f"Your changes have been successfully saved."}, status=200)

        except Products.DoesNotExist:
            return JsonResponse({'error': "Product not found."}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class ExportProductsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        def set_column_widths(worksheet):
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        products = Products.objects.all().values()
        # Create a new workbook and add a worksheet
        jalali_date = current_jalali_date().strftime('%Y-%m-%d')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Products {jalali_date}"
        # Write the header row
        header = ['Group', 'Subgroup', 'Feature', 'Product Code (IR)', 'Product Code (TR)', 'Description (TR)', 
                  'Description (IR)', 'Unit', 'Secondary Unit', 'Weight', 'Currency', 'Price']
        for col_num, column_title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
            cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'),
                                                 bottom=openpyxl.styles.Side(style='medium'),
                                                 left=openpyxl.styles.Side(style='medium'),
                                                 right=openpyxl.styles.Side(style='medium'))
        # Write the data rows
        for row_num, product in enumerate(products, 2):
            row = [product['group'], product['subgroup'], product['feature'], product['product_code_ir'], 
                   product['product_code_tr'], product['description_tr'], product['description_ir'], product['unit'], 
                   product['unit_secondary'], product['weight'], product['currency'], product['price']]
            for col_num, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
        # Apply some styling to the Excel file
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'),
                                                     left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'))
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        # Set the column widths
        set_column_widths(ws)

        # Apply auto filter
        ws.auto_filter.ref = f"A1:L{ws.max_row}"

        # Set the response headers for an Excel file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.read()

        # Encode the content in base64
        base64_content = base64.b64encode(content).decode()

        # Send the content and filename in the JSON response
        return JsonResponse({'filename': f'products({jalali_date}).xlsx', 'content': base64_content})




# endregion

# region Charts

class ChartView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        #start_date = request.POST.get('start_date')
        #end_date = request.POST.get('end_date')
        #product_code = request.POST.get('product_code')
        #start_date = date(1400,4,1)
        #end_date = date(1400,4,30)
        data = Sales.objects.filter(group = "Boya").values('date', 'original_output_value')
        date_list = [obj['date'] for obj in data]
        output_value_list = [obj['original_output_value'] for obj in data]
        response_data = {'date_list': date_list, 'output_value_list': output_value_list}

        return JsonResponse(response_data, safe=False)

class ItemListView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def get(self, request, *args, **kwargs):
        product_codes = Products.objects.values_list('product_code_ir', flat=True)
        return JsonResponse(list(product_codes), safe=False)
    
    def post(self, request, *args, **kwargs):
        # Get the product_title from the POST data

        data = json.loads(request.body)
        product_code = data.get('product_code')

        # Filter Sales by the product_title
        data = Sales.objects.filter(product_code=product_code).values('date', 'original_output_value')
        product_name = Products.objects.filter(product_code_ir=product_code).values('description_ir')
        # Get the original_output_value of each sale
        date_list = [obj['date'] for obj in data]
        output_value_list = [obj['original_output_value'] for obj in data]
        product_name = [obj["product_title"] for obj in product_name]

        response_data = {'product_name':product_name ,'date_list': date_list, 'output_value_list': output_value_list}

        # Return the list of output_values as a JSON response
        return JsonResponse(response_data, safe=False)

# endregion 



# region Saler
class AddSalerView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            print(data)
            jalali_date = data.get("job_start_date").split("-")
            saler_type = data.get("saler_type")
            try:
                jalali_date = jdatetime.date(int(jalali_date[0]), int(jalali_date[1]), int(jalali_date[2]))
            except ValueError:
                return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
            except IndexError as e:
                return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
            except Exception as e:
                return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
            if the_man_from_future(jalali_date):
                return JsonResponse({'error': "HERE'S THE MAN FROM THE FUTURE TO SAVE US ALL!!!! Job Start Date cannot be future time, please check it :) "}, status=400)
            
            if saler_type == "Active":
                bool_active_saler = True
                bool_passive_saler = False
                experience_rating = calculate_experience_rating(jalali_date)
            else:
                bool_active_saler = False
                bool_passive_saler = True
                experience_rating = calculate_passive_saler_experience_rating(jalali_date)
                
            jalali_current_date = current_jalali_date()
           

            saler = Salers(
                name = data.get("name"),
                job_start_date = jalali_date,
                manager_performance_rating = 1,
                experience_rating = experience_rating,
                monthly_total_sales_rating = 1, #will be calculated!!!!!!!!!!!!!!!!!!!!!!
                receipment_rating = 1, #will be calculated!!!!!!!!!!!!!!!!!!!!!!
                is_active =  True,
                is_active_saler = bool_active_saler,
                is_passive_saler = bool_passive_saler   
            )
            saler.save()
            return JsonResponse({'message': "Saler added successfully"}, status=200)
        except IndexError as e:
            return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD' "}, status=400)
        except ValueError as e:
            return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD' "}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)

class EditSalerView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            print(data)

            old_data = data.get('old_data')
            new_data = data.get('new_data')

            # Check if name is provided
            name = new_data['name']
            if not name:
                traceback.print_exc()
                return JsonResponse({'error': "Missing required parameter: 'name'"}, status=400)
            
            # Get the saler object
            saler = Salers.objects.get(id=old_data['id'])
            print(saler)
            
            if saler.is_active_saler == True and saler.is_active_saler == False:
                # Update other saler fields
                for field in ['name', 'job_start_date', 'manager_performance_rating', 'is_active']:
                    if field == "job_start_date":
                        try:
                            new_date =new_data['job_start_date'].split("-")
                            date = jdatetime.date(int(new_date[0]), int(new_date[1]), int(new_date[2]))
                            saler.experience_rating = calculate_experience_rating(date)
                            print("active: ",saler.experience_rating)
                        except ValueError:
                            return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
                        except IndexError as e:
                            return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
                        except Exception as e:
                            traceback.print_exc()
                            return JsonResponse({'error': str(e)}, status=400)
                    value = new_data[f'{field}']

                    if value is not None and value != '':
                        setattr(saler, field, value)
                    else:
                        traceback.print_exc() 
                        return JsonResponse({'error': "One or more data field is empty!"}, status=400)


                saler.save()
                return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)
            else:
                try:
                    # Update other saler fields
                    for field in ['name', 'job_start_date', 'manager_performance_rating', 'is_active']:
                        if field == "job_start_date":
                            try:
                                new_date =new_data['job_start_date'].split("-")
                                date = jdatetime.date(int(new_date[0]), int(new_date[1]), int(new_date[2]))
                                print("date:",date)
                                saler.experience_rating = calculate_passive_saler_experience_rating(date)
                                print("passive: ",saler.experience_rating)
                            except ValueError:
                                return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
                            except IndexError as e:
                                return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)
                            except Exception as e:
                                return JsonResponse({'error': str(e)}, status=400)
                        value = new_data[f'{field}']

                        if value is not None and value != '':
                            setattr(saler, field, value)
                        else:
                            traceback.print_exc() 
                            return JsonResponse({'error': "One or more data field is empty!"}, status=400)


                    saler.save()
                    return JsonResponse({'message': "Your changes have been successfully saved"}, status=200)
                except Exception as e:
                    traceback.print_exc()
                    return JsonResponse({'error': str(e)}, status=500)


        except Salers.DoesNotExist:
            traceback.print_exc()
            return JsonResponse({'error': "Saler not found"}, status=400)

        except ValueError as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)



class CollapsedSalerView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    
    def get(self, request, *args, **kwargs):
        active_salers = Salers.objects.filter(is_deleted=False, is_active_saler=True)
        print(active_salers)
        active_salers_list = [[saler.id, saler.name, saler.is_active] for saler in active_salers]
        passive_salers = Salers.objects.filter(is_deleted=False, is_passive_saler=True)
        passive_salers_list = [[saler.id, saler.name, saler.is_active] for saler in passive_salers]
        
        return JsonResponse({"active_salers_list": active_salers_list,
                             "passive_salers_list": passive_salers_list}, safe=False)


    
 # Everyday experience rating must be automatically updated !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!   
class SalerCardView(APIView):
    permission_classes = (IsAuthenticated,) 
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        id = data.get('id')
        saler = Salers.objects.get(id=id, is_deleted = False)
        jalali_current_date = current_jalali_date()
        try:
            saler_monthly_ratings = SalerMonthlySaleRating.objects.get(name=saler.name, month = jalali_current_date.month, year= jalali_current_date.year )
            monthly_sale_rating = saler_monthly_ratings.sale_rating
        except SalerMonthlySaleRating.DoesNotExist:
            monthly_sale_rating = 1
        if saler.is_active_saler:
            response_data = {'id': id , 'name': saler.name, 'job_start_date': saler.job_start_date.strftime('%Y-%m-%d'), 'manager_performance_rating': saler.manager_performance_rating,
                            'experience_rating': saler.experience_rating, 'monthly_total_sales_rating': monthly_sale_rating, 'receipment_rating':saler.receipment_rating,
                            'is_active': saler.is_active, "active_or_passive": "Active" }
        else:
            response_data = {'id': id , 'name': saler.name, 'job_start_date': saler.job_start_date.strftime('%Y-%m-%d'), 'manager_performance_rating': saler.manager_performance_rating,
                            'experience_rating': saler.experience_rating, 'monthly_total_sales_rating': monthly_sale_rating, 'receipment_rating':saler.receipment_rating,
                            'is_active': saler.is_active, "active_or_passive": "Passive" }
        # Return the list of output_values as a JSON response
        return JsonResponse(response_data, safe=False)

class SalerTableView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        salers = Salers.objects.filter(is_deleted=False).values()
        saler_list = [[s['id'], s['name'], s['job_start_date'].strftime('%Y-%m-%d'), s['manager_performance_rating'],
                       s['experience_rating'], s['monthly_total_sales_rating'], s['receipment_rating'], s['is_active'], s['is_active_saler'] ,s['is_passive_saler']] for s in salers]
        
        print(saler_list)
        return JsonResponse(saler_list, safe=False)

class DeleteSalerView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        try:    
            data = json.loads(request.body)
            id = data.get('id')
            saler = Salers.objects.get(id=id)
            print("saler: ", saler)
            saler.is_deleted = True
            saler.is_active = False
            saler.save()

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        return JsonResponse({'message': "Saler object has been successfully deleted"}, status=200)





@receiver(pre_save, sender=SalerPerformance)
def update_month_sale_rating(sender, instance, **kwargs):
    # Calculate the sale rating based on the updated sale value
    aggregated_sales = SalerPerformance.objects.filter(
        name=instance.name,
        year=instance.year,
        month=instance.month
    ).aggregate(monthly_sale=Sum('sale'))
    monthly_sale = float(aggregated_sales['monthly_sale'] or 0)
    monthly_sale_rating = calculate_sale_rating(monthly_sale / 10000000)
    saler_rating, created = SalerMonthlySaleRating.objects.get_or_create(
        name=instance.name, 
        year=instance.year,
        month=instance.month
        )
    saler_rating.sale_rating = monthly_sale_rating
    saler_rating.save()
    saler = Salers.objects.get(name=instance.name)
    saler.monthly_total_sales_rating = monthly_sale_rating





# endregion

# region SalerPerformance

@receiver(post_save, sender=Sales)
def update_saler_performance_with_add_sale(sender, instance, created, **kwargs):
    # Get or create the SalerPerformance object
    saler_performance, _ = SalerPerformance.objects.get_or_create(
        name=instance.saler,
        year=instance.date.year,
        month=instance.date.month,
        day=instance.date.day
    )

    if created:
        # Update the sale value for the SalerPerformance object
        saler_performance.sale += float(instance.net_sales)
        saler_performance.bonus += float(instance.bonus)
    else:
        # Check which fields have been updated
        dirty_fields = instance.get_dirty_fields()

        # If any of the saler, year, month, or day fields are updated, find the previous SalerPerformance instance, subtract the old values, and update the new SalerPerformance instance
        if any(field in dirty_fields for field in ['saler', 'date']):
            old_saler = dirty_fields.get('saler', instance.saler)
            old_date = dirty_fields.get('date', instance.date)
            old_saler_performance = SalerPerformance.objects.get(
                name=old_saler,
                year=old_date.year,
                month=old_date.month,
                day=old_date.day
            )

            # Subtract old values from the old SalerPerformance instance
            old_saler_performance.sale -= dirty_fields.get('net_sales', instance.net_sales)
            old_saler_performance.bonus -= dirty_fields.get('bonus', instance.bonus)
            old_saler_performance.save()

            # Update the new SalerPerformance instance
            saler_performance.sale += instance.net_sales
            saler_performance.bonus += instance.bonus

        # Update the corresponding attributes of the SalerPerformance instance based on the updated fields
        else:
            if 'net_sales' in dirty_fields:
                saler_performance.sale += float(instance.net_sales) - dirty_fields['net_sales']
            if 'bonus' in dirty_fields:
                saler_performance.bonus += float(instance.bonus) - dirty_fields['bonus']

    saler_performance.save()


@receiver(post_delete, sender=Sales)
def update_saler_performance_with_delete_sale(sender, instance, **kwargs):
    # Get the corresponding SalerPerformance object for the sale
    performance, created = SalerPerformance.objects.get_or_create(
        name=instance.saler, 
        year=instance.date.year,
        month=instance.date.month,
        day=instance.date.day
        )

    # Subtract the net sale amount from the sale field
    performance.sale -= instance.net_sales
    performance.bonus -= instance.bonus
    performance.save()


class SalerPerformanceView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        saler_performances = SalerPerformance.objects.values().all()
        saler_performance_list = [[performance['name'], performance['year'], performance['month'],
                                   performance['day'], performance['sale'], performance['bonus']]
                                  for performance in saler_performances]
        return JsonResponse(saler_performance_list, safe=False)



# endregion

# region SaleSummary

@receiver(post_save, sender=Sales)
def update_sale_summary_with_add_sale(sender, instance, created, **kwargs):
    find_month = instance.date.month
    find_year = instance.date.year
    find_day = instance.date.day
    sale_summary, _ = SaleSummary.objects.get_or_create(
        date=jdatetime.date(int(find_year), int(find_month), int(find_day)),
        year=find_year,
        month=find_month,
        day=find_day
    )

    if created:
        # Add the values of all relevant fields to the corresponding attributes of the SaleSummary instance
        sale_summary.sale += instance.net_sales
        sale_summary.dollar_sepidar_sale += instance.dollar_sepidar
        sale_summary.dollar_sale += instance.dollar
        sale_summary.kg_sale += instance.kg
        sale_summary.save()
    else:
        # Check which fields have been updated
        dirty_fields = instance.get_dirty_fields()

        # If the date field is updated, find the previous SaleSummary instance, subtract the old values, and update the new SaleSummary instance
        if 'date' in dirty_fields:
            old_date = dirty_fields['date']
            old_sale_summary = SaleSummary.objects.get(
                year=old_date.year,
                month=old_date.month,
                day=old_date.day
            )

            # Subtract old values from the old SaleSummary instance
            old_sale_summary.sale -= dirty_fields.get('net_sales', instance.net_sales)
            old_sale_summary.dollar_sepidar_sale -= dirty_fields.get('dollar_sepidar', instance.dollar_sepidar)
            old_sale_summary.dollar_sale -= dirty_fields.get('dollar', instance.dollar)
            old_sale_summary.kg_sale -= dirty_fields.get('kg', instance.kg)
            old_sale_summary.save()

            # Update the new SaleSummary instance
            sale_summary.sale += instance.net_sales
            sale_summary.dollar_sepidar_sale += instance.dollar_sepidar
            sale_summary.dollar_sale += instance.dollar
            sale_summary.kg_sale += instance.kg

        # Update the corresponding attributes of the SaleSummary instance based on the updated fields
        else:
            if 'net_sales' in dirty_fields:
                sale_summary.sale += float(instance.net_sales) - dirty_fields['net_sales']
            if 'dollar_sepidar' in dirty_fields:
                sale_summary.dollar_sepidar_sale += float(instance.dollar_sepidar) - dirty_fields['dollar_sepidar']
            if 'dollar' in dirty_fields:
                sale_summary.dollar_sale += float(instance.dollar) - dirty_fields['dollar']
            if 'kg' in dirty_fields:
                sale_summary.kg_sale += float(instance.kg) - dirty_fields['kg']

        sale_summary.save()

@receiver(post_delete, sender=Sales)
def update_sale_summary_with_delete_sale(sender, instance, **kwargs):
    # Get or create the SaleSummary object
    find_month = instance.date.month
    find_year = instance.date.year
    find_day = instance.date.day

    sale_summary = SaleSummary.objects.get(
        date=jdatetime.date(int(find_year), int(find_month), int(find_day))
    )

    # Update the sale value for the SaleSummary object
    sale_summary.sale -= instance.net_sales
    sale_summary.dollar_sepidar_sale -= instance.dollar_sepidar
    sale_summary.dollar_sale -= instance.dollar
    sale_summary.kg_sale -= instance.kg
    sale_summary.save()

class SalesReportView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)

        report_type = data.get('report_type')
        start_date =  data.get('start_date').split("-")
        end_date =  data.get('end_date').split("-")
       

        if report_type == 'daily':
            start_date = jdatetime.date(int(start_date[0]), int(start_date[1]), int(start_date[2]))
            end_date = jdatetime.date(int(end_date[0]), int(end_date[1]), int(end_date[2]))
            data = SaleSummary.objects.filter(date__range = [start_date, end_date]).values('date').annotate(total_sales=Sum('sale')).order_by('date')
            sales_report_list = [[d['date'].strftime('%Y-%m-%d'), d['total_sales']] for d in data]

        elif report_type == 'monthly':
            start_year, start_month = int(start_date[0]), int(start_date[1])
            end_year, end_month = int(end_date[0]), int(end_date[1])
            sales_report_list = []
            
            while start_year < end_year or (start_year == end_year and start_month < end_month):
                # Get the data for the current month
                data = SaleSummary.objects.filter(year=start_year, month=start_month).values('year', 'month').annotate(total_sales=Sum('sale'))

                # Format the data for the current month and add it to the sales report list
                sales_report_list.extend([[f"{jdatetime.date(start_year, start_month, 1).strftime('%Y-%m-%d')}", d['total_sales']] for d in data])

                # Increment the month and year
                start_month += 1
                if start_month > 12:
                    start_month = 1
                    start_year += 1

            # Sort the sales report list by date
            sales_report_list.sort()
        elif report_type == 'yearly':
            start_year  = int(start_date[0])
            end_year  = int(end_date[0])
            sales_report_list = []
            
            while start_year < end_year :
                # Get the data for the current month
                data = SaleSummary.objects.filter(year=start_year).values('year').annotate(total_sales=Sum('sale'))

                # Format the data for the current month and add it to the sales report list
                sales_report_list.extend([[f"{jdatetime.date(start_year, 1, 1).strftime('%Y-%m-%d')}", d['total_sales']] for d in data])

                # Increment the month and year
                start_year += 1

            # Sort the sales report list by date
            sales_report_list.sort()

        else:
            data = []
        return JsonResponse(sales_report_list, safe=False)




# endregion

# region MonthlyProductSales

@receiver(post_save, sender=Sales)
def update_monthly_product_sales_with_add_sale(sender, instance, created, **kwargs):
    # Get or create the MonthlyProductSales object
    monthly_sale, _ = MonthlyProductSales.objects.get_or_create(
        product_code=instance.product_code,
        year=instance.date.year,
        month=instance.date.month
    )
    monthly_sale.date = instance.date
    monthly_sale.product_name = instance.product_name

    if created:
        # Update the MonthlyProductSales instance with the new sales information
        monthly_sale.piece += instance.original_output_value
        monthly_sale.sale += instance.net_sales
    else:
        # Check which fields have been updated
        dirty_fields = instance.get_dirty_fields()

        # If any of the product_code, year, or month fields are updated, find the previous MonthlyProductSales instance, subtract the old values, and update the new MonthlyProductSales instance
        if any(field in dirty_fields for field in ['product_code', 'date']):
            old_product_code = dirty_fields.get('product_code', instance.product_code)
            old_date = dirty_fields.get('date', instance.date)
            old_monthly_sale = MonthlyProductSales.objects.get(
                product_code=old_product_code,
                year=old_date.year,
                month=old_date.month
            )

            # Subtract old values from the old MonthlyProductSales instance
            old_monthly_sale.piece -= dirty_fields.get('original_output_value', instance.original_output_value)
            old_monthly_sale.sale -= dirty_fields.get('net_sales', instance.net_sales)
            old_monthly_sale.save()

            # Update the new MonthlyProductSales instance
            monthly_sale.piece += instance.original_output_value
            monthly_sale.sale += instance.net_sales

        # Update the corresponding attributes of the MonthlyProductSales instance based on the updated fields
        else:
            if 'original_output_value' in dirty_fields:
                monthly_sale.piece += float(instance.original_output_value) - dirty_fields['original_output_value']
            if 'net_sales' in dirty_fields:
                monthly_sale.sale += float(instance.net_sales) - dirty_fields['net_sales']

    monthly_sale.save()



@receiver(post_delete, sender=Sales)
def update_monthly_product_sales_with_delete_sale(sender, instance, **kwargs):

    monthly_sale = MonthlyProductSales.objects.get(
        product_code=instance.product_code,
        year=instance.date.year,
        month= instance.date.month
        )

    monthly_sale.product_name = instance.product_name,
    monthly_sale.piece -= instance.original_output_value
    monthly_sale.sale -= instance.net_sales
    monthly_sale.save()

# endregion


# region Dashboard #!DASHBOARD PAGE START

# region Customer Performance

@receiver(post_save, sender=Sales)
def update_customer_performance_with_add_sale(sender, instance, created, **kwargs):
    # Get or create the CustomerPerformance object
    find_month = instance.date.month
    find_year = instance.date.year
    customer_performance, _ = CustomerPerformance.objects.get_or_create(
         year=find_year, month=find_month, customer_code=instance.customer_code
    )

    # Update the corresponding attributes of the CustomerPerformance instance based on the updated fields
    customer_performance.customer_name = instance.name
    customer_performance.customer_area = instance.area

    if created:
        # Update the sale value for the CustomerPerformance object
        customer_performance.sale += instance.net_sales
        customer_performance.sale_amount += instance.original_output_value
        customer_performance.dollar += instance.dollar
        customer_performance.dollar_sepidar += instance.dollar_sepidar
    else:
        # Check which fields have been updated
        dirty_fields = instance.get_dirty_fields()

        # If any of the year, month, or customer_code fields are updated, find the previous CustomerPerformance instance, subtract the old values, and update the new CustomerPerformance instance
        if any(field in dirty_fields for field in ['date', 'customer_code']):
            old_date = dirty_fields.get('date', instance.date)
            old_customer_code = dirty_fields.get('customer_code', instance.customer_code)
            old_customer_performance = CustomerPerformance.objects.get(
                year=old_date.year, month=old_date.month, customer_code=old_customer_code
            )

            # Subtract old values from the old CustomerPerformance instance
            old_customer_performance.sale -= dirty_fields.get('net_sales', instance.net_sales)
            old_customer_performance.sale_amount -= dirty_fields.get('original_output_value', instance.original_output_value)
            old_customer_performance.dollar -= dirty_fields.get('dollar', instance.dollar)
            old_customer_performance.dollar_sepidar -= dirty_fields.get('dollar_sepidar', instance.dollar_sepidar)
            old_customer_performance.save()

            # Update the new CustomerPerformance instance
            customer_performance.sale += instance.net_sales
            customer_performance.sale_amount += instance.original_output_value
            customer_performance.dollar += instance.dollar
            customer_performance.dollar_sepidar += instance.dollar_sepidar

        # Update the corresponding attributes of the CustomerPerformance instance based on the updated fields
        else:
            if 'net_sales' in dirty_fields:
                customer_performance.sale += float(instance.net_sales) - dirty_fields['net_sales']
            if 'original_output_value' in dirty_fields:
                customer_performance.sale_amount += float(instance.original_output_value) - dirty_fields['original_output_value']
            if 'dollar' in dirty_fields:
                customer_performance.dollar += float(instance.dollar) - dirty_fields['dollar']
            if 'dollar_sepidar' in dirty_fields:
                customer_performance.dollar_sepidar += float(instance.dollar_sepidar) - dirty_fields['dollar_sepidar']

    customer_performance.save()



@receiver(post_delete, sender=Sales)
def update_customer_performance_with_delete_sale(sender, instance, **kwargs):
    # Get or create the CustomerPerformance object
    find_month = instance.date.month
    find_year = instance.date.year
    customer_performance = CustomerPerformance.objects.get(
            year= find_year, month = find_month, customer_code = instance.customer_code
    )

    # Update the sale value for the CustomerPerformance object
    customer_performance.customer_name = instance.name
    customer_performance.customer_area = instance.area
    customer_performance.sale -= instance.net_sales
    customer_performance.sale_amount -= instance.original_output_value
    customer_performance.dollar -= instance.dollar
    customer_performance.dollar_sepidar -= instance.dollar_sepidar
    customer_performance.save()

class TopCustomersView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)

        report_type = data.get('report_type')
        if report_type == 'monthly':
            top_customers_list = []
            
            # Get the current month and year using jdatetime library
            date_month= current_jalali_date().month
            date_year= current_jalali_date().year
            
            # Get the data for the current month
            top_5_customer_data = CustomerPerformance.objects.filter(month=date_month, year=date_year).order_by('-sale')[:5]
            
            # Calculate the total sales for the current month
            total_sales = CustomerPerformance.objects.filter(month=date_month, year=date_year).aggregate(total_sales=Sum('sale'))['total_sales']
            if total_sales is not None:
                # Calculate the sales data for the top 5 customers and others
                top_customers_list = [[d.customer_name, d.sale] for d in top_5_customer_data]
                top_customers_sale_sum = [d[1] for d in top_customers_list]
                top_5_customer_total_sale = sum(top_customers_sale_sum)
                others_sales = total_sales - top_5_customer_total_sale

                # Create a list of sales data for the top 5 customers and others for the pie chart
                top_customers_pie_chart = [[d[0], (d[1]/total_sales)*100] for d in top_customers_list]
                top_customers_pie_chart.append(["others", (others_sales/total_sales*100)])
            else:
                # Handle the case when there is no sales data available
                top_customers_pie_chart = [["No data available", 100]]
        
        elif report_type == 'yearly':
            top_customers_list = []
            
            # Get the current year using jdatetime library
            date= current_jalali_date().year
            
            # Get the data for the current year
            top_5_customer_data = CustomerPerformance.objects.filter(year=date).order_by('-sale')[:5]
            
            # Calculate the total sales for the current year
            total_sales = CustomerPerformance.objects.filter(year=date).aggregate(total_sales=Sum('sale'))['total_sales']
            if total_sales is not None:
                # Calculate the sales data for the top 5 customers and others
                top_customers_list = [[d.customer_name, d.sale] for d in top_5_customer_data]
                top_customers_sale_sum = [d[1] for d in top_customers_list ]
                top_5_customer_total_sale = sum(top_customers_sale_sum)
                others_sales = total_sales - top_5_customer_total_sale
                
                # Create a list of sales data for the top 5 customers and others for the pie chart
                top_customers_pie_chart = [[d[0], (d[1]/total_sales)*100] for d in top_customers_list]
                top_customers_pie_chart.append(["others", (others_sales/total_sales*100)])
            else:
                # Handle the case when there is no sales data available
                top_customers_pie_chart = [["No data available", 100]]
        
        return JsonResponse({"top_customers_list": top_customers_list,"top_customers_pie_chart": top_customers_pie_chart}, safe=False)

# endregion


# region Product Performance

@receiver(post_save, sender=Sales)
def update_product_performance_with_add_sale(sender, instance, created, **kwargs):
    # Get or create the ProductPerformance object
    find_month = instance.date.month
    find_year = instance.date.year
    product_performance, _ = ProductPerformance.objects.get_or_create(
         year=find_year, month=find_month, product_code=instance.product_code
    )

    # Update the corresponding attributes of the ProductPerformance instance based on the updated fields
    product_performance.product_name = instance.product_name

    if created:
        # Update the sale value for the ProductPerformance object
        product_performance.sale_amount += instance.original_output_value
        product_performance.sale += instance.net_sales
    else:
        # Check which fields have been updated
        dirty_fields = instance.get_dirty_fields()

        # If any of the year, month, or product_code fields are updated, find the previous ProductPerformance instance, subtract the old values, and update the new ProductPerformance instance
        if any(field in dirty_fields for field in ['date', 'product_code']):
            old_date = dirty_fields.get('date', instance.date)
            old_product_code = dirty_fields.get('product_code', instance.product_code)
            old_product_performance = ProductPerformance.objects.get(
                year=old_date.year, month=old_date.month, product_code=old_product_code
            )

            # Subtract old values from the old ProductPerformance instance
            old_product_performance.sale_amount -= dirty_fields.get('original_output_value', instance.original_output_value)
            old_product_performance.sale -= dirty_fields.get('net_sales', instance.net_sales)
            old_product_performance.save()

            # Update the new ProductPerformance instance
            product_performance.sale_amount += instance.original_output_value
            product_performance.sale += instance.net_sales

        # Update the corresponding attributes of the ProductPerformance instance based on the updated fields
        else:
            if 'original_output_value' in dirty_fields:
                product_performance.sale_amount += float(instance.original_output_value) - dirty_fields['original_output_value']
            if 'net_sales' in dirty_fields:
                product_performance.sale += float(instance.net_sales) - dirty_fields['net_sales']

    product_performance.save()



@receiver(post_delete, sender=Sales)
def update_product_performance_with_delete_sale(sender, instance, **kwargs):
    # Get or create the ProductPerformance object
    find_month = instance.date.month
    find_year = instance.date.year
    product_performance = ProductPerformance.objects.get(
            year= find_year, month = find_month, product_code = instance.product_code
    )

    # Update the sale value for the ProductPerformance object
    product_performance.product_name = instance.product_name
    product_performance.sale_amount -= instance.original_output_value
    product_performance.sale -= instance.net_sales
    product_performance.save()

class TopProductsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)

        report_type = data.get('report_type')
        if report_type == 'monthly':
            top_products_list = []
            
            # Get the current month and year using jdatetime library
            date_month= current_jalali_date().month
            date_year= current_jalali_date().year
            
            # Get the data for the current month
            top_5_product_data = ProductPerformance.objects.filter(month=date_month, year=date_year).order_by('-sale')[:5]
            
            # Calculate the total sales for the current month
            total_sales = ProductPerformance.objects.filter(month=date_month, year=date_year).aggregate(total_sales=Sum('sale'))['total_sales']
            if total_sales is not None:
                # Calculate the sales data for the top 5 products and others
                top_products_list = [[d.product_name, d.sale] for d in top_5_product_data]
                top_products_sale_sum = [d[1] for d in top_products_list ]
                top_5_product_total_sale = sum(top_products_sale_sum)
                others_sales = total_sales - top_5_product_total_sale

                # Create a list of sales data for the top 5 products and others for the pie chart
                top_products_pie_chart = [[d[0], (d[1]/total_sales)*100] for d in top_products_list]
                top_products_pie_chart.append(["others", (others_sales/total_sales*100)])
            else:
                # Handle the case when there is no sales data available
                top_products_pie_chart = [["No data available", 100]]
        
        elif report_type == 'yearly':
            top_products_list = []
            
            # Get the current year using jdatetime library
            date= current_jalali_date().year
            
            # Get the data for the current year
            top_5_product_data = ProductPerformance.objects.filter(year=date).order_by('-sale')[:5]
            
            # Calculate the total sales for the current year
            total_sales = ProductPerformance.objects.filter(year=date).aggregate(total_sales=Sum('sale'))['total_sales']
            if total_sales is not None:
                # Calculate the sales data for the top 5 products and others
                top_products_list = [[d.product_name, d.sale] for d in top_5_product_data]
                top_products_sale_sum = [d[1] for d in top_products_list ]
                top_5_product_total_sale = sum(top_products_sale_sum)
                others_sales = total_sales - top_5_product_total_sale

                # Create a list of sales data for the top 5 products and others for the pie chart
                top_products_pie_chart = [[d[0], (d[1]/total_sales)*100] for d in top_products_list]
                top_products_pie_chart.append(["others", (others_sales/total_sales*100)])
            else:
                # Handle the case when there is no sales data available
                top_products_pie_chart = [["No data available", 100]]
        
        return JsonResponse({"top_products_list": top_products_list,"top_products_pie_chart": top_products_pie_chart}, safe=False)

# endregion

# region Current Exchange Rate

class ExchangeRateAPIView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def get(self, request):
        try:
            exchange_rate = get_exchange_rate()
            response_data = exchange_rate
        except Exception as e:
            response_data = {
                "error": "There is an error at Current IRR Exchange Rate. Please contact developer to solve it",
            }

        return JsonResponse(response_data,safe=False )

# endregion

# region Daily Report

class SalerDataView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        jalali_date_now = current_jalali_date()
        jalali_date_now_str = jalali_date_now.strftime('%Y-%m-%d')

        daily_sales = SalerPerformance.objects.filter(
            year=jalali_date_now.year,
            month=jalali_date_now.month,
            day=jalali_date_now.day
        ).values('name', 'sale')

        monthly_sales = SalerPerformance.objects.filter(
            year=jalali_date_now.year,
            month=jalali_date_now.month
        ).values('name').annotate(monthly_sale=Sum('sale'))

        yearly_sales = SalerPerformance.objects.filter(
            year=jalali_date_now.year
        ).values('name').annotate(yearly_sale=Sum('sale'))

        # Combine the data into a single list
        combined_data = []
        if not daily_sales:
            # If no sales for the current day, append a dictionary with name and daily sale values set to zero
            for monthly_sale in monthly_sales:
                name = monthly_sale['name']
                yearly_sale = next((item['yearly_sale'] for item in yearly_sales if item['name'] == name), 0)
                try:
                    saler = Salers.objects.get(name=name)
                    is_active = saler.is_active
                except:
                    is_active = "Left"
                combined_data.append([
                    name,
                    is_active,
                    0,  # Set daily sale value to zero
                    monthly_sale['monthly_sale'] / 10,
                    yearly_sale / 10
                ])
        else:
            for daily_sale in daily_sales:
                name = daily_sale['name']
                try:
                    saler = Salers.objects.get(name=name)
                    is_active = saler.is_active
                except:
                    is_active = "Left"
                monthly_sale = next((item['monthly_sale'] for item in monthly_sales if item['name'] == name), 0)
                yearly_sale = next((item['yearly_sale'] for item in yearly_sales if item['name'] == name), 0)

                combined_data.append([
                    name,
                    is_active,
                    daily_sale['sale'] / 10,
                    monthly_sale / 10,
                    yearly_sale / 10
                ])

        response_data = {"jalali_date": jalali_date_now_str, "sales_data": combined_data}

        return JsonResponse(response_data, safe=False)
        

class TotalDataView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def get(self, request, *args, **kwargs):
        jalali_date_now = current_jalali_date()
        jalali_date_now_str = jalali_date_now.strftime('%Y-%m-%d')
        
        daily_sales = SaleSummary.objects.filter(
            year=jalali_date_now.year,
            month=jalali_date_now.month,
            day=jalali_date_now.day
        ).values('sale', 'dollar_sepidar_sale', 'dollar_sale', 'kg_sale')
        daily_sales_array = list(daily_sales.values_list('sale', 'dollar_sepidar_sale', 'dollar_sale', 'kg_sale')[0]) if daily_sales.exists() else [0, 0, 0, 0]
        # Divide each value in daily_sales_array by 10
        for i in range(len(daily_sales_array)):
            daily_sales_array[i] /= 10

        monthly_sales = SaleSummary.objects.filter(
            year=jalali_date_now.year,
            month=jalali_date_now.month
        ).annotate(monthly_sale=Sum('sale'), monthly_dollar_sepidar_sale=Sum('dollar_sepidar_sale'), monthly_dollar_sale=Sum('dollar_sale'), monthly_kg_sale=Sum('kg_sale') )
        monthly_sales_array = list(monthly_sales.values_list('monthly_sale', 'monthly_dollar_sepidar_sale', 'monthly_dollar_sale', 'monthly_kg_sale')[0]) if monthly_sales.exists() else [0, 0, 0, 0]
        # Divide each value in daily_sales_array by 10
        for i in range(len(monthly_sales_array)):
            monthly_sales_array[i] /= 10


        yearly_sales = SaleSummary.objects.filter(
            year=jalali_date_now.year
        ).annotate(yearly_sale=Sum('sale'), yearly_dollar_sepidar_sale=Sum('dollar_sepidar_sale'), yearly_dollar_sale=Sum('dollar_sale'), yearly_kg_sale=Sum('kg_sale') )
        yearly_sales_array = list(yearly_sales.values_list('yearly_sale', 'yearly_dollar_sepidar_sale', 'yearly_dollar_sale', 'yearly_kg_sale')[0]) if yearly_sales.exists() else [0, 0, 0, 0]
        # Divide each value in daily_sales_array by 10
        for i in range(len(yearly_sales_array)):
            yearly_sales_array[i] /= 10
        

        response_data = { "jalali_date" : jalali_date_now_str, "daily_sales" : daily_sales_array, "monthly_sales" : monthly_sales_array, "yearly_sales" : yearly_sales_array }

        # Combine the data into a single list
        
        
        return JsonResponse(response_data, safe=False)
        

class TotalDataByMonthlyView(View):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def get(self, request, *args, **kwargs):
        jalali_date_now = current_jalali_date()
        jalali_date_now_str = jalali_date_now.strftime('%Y-%m-%d')
        
        monthly_sales = SaleSummary.objects.filter(
        year=jalali_date_now.year,
        month__lte=jalali_date_now.month
        ).values('month').annotate(monthly_sale=Sum('sale'), monthly_dollar_sepidar_sale=Sum('dollar_sepidar_sale'), monthly_dollar_sale=Sum('dollar_sale'), monthly_kg_sale=Sum('kg_sale'))

        monthly_sales_dict = {}
        for monthly_sale_obj in monthly_sales:
            monthly_sales_dict[monthly_sale_obj['month']] = {
                'monthly_sale': monthly_sale_obj['monthly_sale'] / 10,
                'monthly_dollar_sepidar_sale': monthly_sale_obj['monthly_dollar_sepidar_sale'],
                'monthly_dollar_sale': monthly_sale_obj['monthly_dollar_sale'],
                'monthly_kg_sale': monthly_sale_obj['monthly_kg_sale']
            }

        monthly_sales_data = []
        for i in range(1, 13):
            monthly_sales_obj = monthly_sales_dict.get(i)
            if monthly_sales_obj:
                dollar_sale_per_kg = monthly_sales_obj['monthly_dollar_sale'] / monthly_sales_obj['monthly_kg_sale'] if monthly_sales_obj['monthly_kg_sale'] != 0 else 0
                monthly_sales_data.append([
                    monthly_sales_obj['monthly_sale'],
                    monthly_sales_obj['monthly_dollar_sepidar_sale'],
                    monthly_sales_obj['monthly_dollar_sale'],
                    monthly_sales_obj['monthly_kg_sale'],
                    dollar_sale_per_kg
                ])
            elif i <= jalali_date_now.month:
                monthly_sales_data.append([0, 0, 0, 0, 0])

        monthly_sales_array = monthly_sales_data if len(monthly_sales_data) == 12 else monthly_sales_data + [[0, 0, 0, 0, 0]] * (12 - len(monthly_sales_data))
        return JsonResponse(monthly_sales_array, safe=False)


# endregion

# region Customer Area 

class CustomerAreaPieChartView(View):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        
        report_type = data.get('report_type')
        date_month= current_jalali_date().month
        date_year= current_jalali_date().year
        
        table_data = []
        if report_type == 'monthly':
            data = CustomerPerformance.objects.filter(year=date_year, month=date_month).values('customer_area').annotate(total_dollar=Sum('dollar'))
        else:  # default is 'yearly'
            data = CustomerPerformance.objects.filter(year=date_year).values('customer_area').annotate(total_dollar=Sum('dollar'))

        
        total_dollar = sum([item['total_dollar'] for item in data])
        if total_dollar is not None and total_dollar != 0:
            table_data = [[item['customer_area'], item['total_dollar']] for item in data]
            chart_data_percent = [[item['customer_area'], item['total_dollar'] / total_dollar * 100] for item in data]
            
        else:
            # Handle the case when there is no sales data available
            chart_data_percent = [["No data available", 100]]
        
        return JsonResponse({"table_data": table_data, "chart_data_percent": chart_data_percent}, safe=False)

# endregion

# endregion #!DASHBOARD PAGE END

# region ROP
@receiver(post_save, sender=Warehouse)
def create_rop_for_warehouse(sender, instance, created, **kwargs):
    if created:
        try:
            product = Products.objects.get(product_code_ir = instance.product_code)
            
            rop, created = ROP.objects.get_or_create(
            group = product.group,
            subgroup = product.subgroup,
            feature = product.feature,
            new_or_old_product = 0,
            related = None,
            origin = None,
            product_code_ir = instance.product_code,
            product_code_tr = product.product_code_tr,
            dont_order_again = None,
            description_tr = product.description_tr,
            description_ir = product.description_ir,
            unit = product.unit,
            weight = product.weight,
            unit_secondary = product.unit_secondary,
            price = product.price,
            #color_making_room_1400 = None,
            avarage_previous_year = None,
            month_1 = None,
            month_2 = None,
            month_3 = None,
            month_4 = None,
            month_5 = None,
            month_6 = None,
            month_7 = None,
            month_8 = None,
            month_9 = None,
            month_10 = None,
            month_11 = None,
            month_12 = None,
            total_sale = None,
            warehouse = None,
            goods_on_the_road = None,
            total_stock_all = None,
            total_month_stock = None,
            standart_deviation = None,
            lead_time = None,
            product_coverage_percentage = None,
            demand_status = None,
            safety_stock = None,
            rop = None,
            monthly_mean = None,
            new_party = None,
            cycle_service_level = None,
            total_stock = None,
            need_prodcuts = None,
            over_stock = None,
            calculated_need = None,
            calculated_max_stock = None,
            calculated_min_stock = None,
            )

            rop.save()

        except Exception as e:
            return JsonResponse({'error': str(e)})

@receiver(post_save, sender=Sales)
def update_rop_for_sales_add_or_edit(sender, instance, created, **kwargs):
        try:
            product = Products.objects.get(product_code_ir = instance.product_code)
            warehouse = Warehouse.objects.get(product_code=instance.product_code)
            product_performance = ProductPerformance.objects.filter(product_code =instance.product_code )
            jalali_date_now = current_jalali_date()
            jalali_date_now_year = int(jalali_date_now.year)
            jalali_date_now_month = int(jalali_date_now.month)
            jalali_date_previous_year = jalali_date_now_year-1
            product_performance_previous_year = product_performance.filter(year=jalali_date_previous_year)
            # calculate the average sale amount
            average_sale_amount_previous_year = product_performance_previous_year.aggregate(avg_sale=Avg('sale_amount'))['avg_sale']
            
            rop = ROP.objects.get(
                product_code_ir = instance.product_code
                )
            
            # rop.group = product.group,
            # rop.subgroup = product.subgroup,
            # rop.feature = product.feature,
            # rop.new_or_old_product = 0,
            # rop.related = None,
            # rop.origin = None,
            # rop.product_code_ir = instance.product_code,
            # rop.product_code_tr = product.product_code_tr,
            # rop.dont_order_again = None,
            # rop.description_tr = product.description_tr,
            # rop.description_ir = product.description_ir,
            # rop.unit = product.unit,
            # rop.weight = product.weight,
            # rop.unit_secondary = product.unit_secondary,
            # rop.price = product.price,
            # #rop.color_making_room_1400 = None,
            rop.avarage_previous_year = average_sale_amount_previous_year
            
            # amount sales from month_1 to month_12
            for month_number in range(1, 13):
                # Filter the `ProductPerformance` objects by the current year and month.
                product_performance_current_month = product_performance.filter(year=jalali_date_now_year, month=month_number)

                # Calculate the total sale amount for the current month.
                total_sale_current_month = product_performance_current_month.aggregate(total_sale=Coalesce(Sum('sale_amount', output_field=models.FloatField()), float(0)))['total_sale']


                # Set the total sale amount for the current month for the corresponding `ROP` month field.
                setattr(rop, f'month_{month_number}', total_sale_current_month)
            rop.total_sale = product_performance.filter(year=jalali_date_now_year).aggregate(total_sale=Coalesce(Sum('sale_amount'), float(0)))['total_sale']
            rop.warehouse = warehouse.stock
            rop.goods_on_the_road = float(0) #! goods on road values will be updated!!!!!!!!!!!!!!!!!!!!!!!!!!!!
            rop.total_stock_all = rop.warehouse+ rop.goods_on_the_road
            rop.monthly_mean = rop.total_sale/int(jalali_date_now.month)
            try: 
                rop.total_month_stock = rop.total_stock_all/rop.monthly_mean
            except Exception as e:
                rop.total_month_stock = 0
            
            # Get the values of `month_1` to the current month.
            values = [getattr(rop, f'month_{i}', 0) for i in range(1, jalali_date_now_month + 1)]
            # Calculate the standard deviation.
            if len(values) > 0:
                standard_deviation = statistics.stdev(values)
            else:
                standard_deviation = 0

            # Set the `standart_deviation` field of the `ROP` instance.
            rop.standart_deviation = standard_deviation
            
            rop.lead_time = 2 #! will be given by user
            rop.product_coverage_percentage = 95 #! will be given by user
            rop.demand_status =  rop.standart_deviation * (rop.lead_time)**0.5

            #safety stock
            try:
                result = norm.ppf(rop.product_coverage_percentage) * rop.demand_status
                result_rounded_up = math.ceil(result)           
                rop.safety_stock = result_rounded_up
            except:
                rop.safety_stock = float(0)

            rop.rop = (rop.lead_time * rop.monthly_mean) + rop.safety_stock 
            #rop.monthly_mean = rop.total_sale/int(jalali_date_now.month)
            rop.new_party = rop.lead_time * rop.monthly_mean
            
            # cycle_service_level
            try:
                # Calculate the PDF of the normal distribution.
                pdf_value = norm.pdf(rop.rop, loc=rop.monthly_mean, scale=rop.demand_status)

                # Set the result to the PDF.
                rop.cycle_service_level = pdf_value
            except:
                # If an error occurs, set the result to 0.
                rop.cycle_service_level = 0
            rop.total_stock = rop.total_stock_all
            
            # need_products
            if rop.rop >= rop.total_stock:
                rop.need_prodcuts = rop.rop + rop.new_party - rop.total_stock
            else:
                 rop.need_prodcuts = 0

            # over stock
            if rop.total_stock_all > (1.2*(rop.safety_stock + rop.new_party)):  #! Stock Over Factor will be declared and it will produced by user
                rop.over_stock = 1
            else:
                rop.over_stock = 0
            
            rop.calculated_need = rop.need_prodcuts #! previous year is needed??????
            
            # calculated max stock #! previous year is needed??????
            try:
                rop.calculated_max_stock = (rop.rop + rop.new_party)/rop.monthly_mean
            except:
                rop.calculated_max_stock = 0

            # calculated min stock #! previous year is needed??????
            try:
                rop.calculated_min_stock = rop.rop /rop.monthly_mean
            except:
                rop.calculated_min_stock = 0
            rop.save()
        except Products.DoesNotExist:
            pass
        except Exception as e:
            return JsonResponse({'error': str(e)})

@receiver(post_delete, sender=Sales)
def update_rop_for_sales_delete(sender, instance, **kwargs):
        try:
            product = Products.objects.get(product_code_ir = instance.product_code)
            warehouse = Warehouse.objects.get(product_code=instance.product_code)
            product_performance = ProductPerformance.objects.filter(product_code =instance.product_code )
            jalali_date_now = current_jalali_date()
            jalali_date_now_year = int(jalali_date_now.year)
            jalali_date_now_month = int(jalali_date_now.month)
            jalali_date_previous_year = jalali_date_now_year-1
            product_performance_previous_year = product_performance.filter(year=jalali_date_previous_year)
            # calculate the average sale amount
            average_sale_amount_previous_year = product_performance_previous_year.aggregate(avg_sale=Avg('sale_amount'))['avg_sale']
            
            rop = ROP.objects.get(
                product_code_ir = instance.product_code
                )
            
            # rop.group = product.group,
            # rop.subgroup = product.subgroup,
            # rop.feature = product.feature,
            # rop.new_or_old_product = 0,
            # rop.related = None,
            # rop.origin = None,
            # rop.product_code_ir = instance.product_code,
            # rop.product_code_tr = product.product_code_tr,
            # rop.dont_order_again = None,
            # rop.description_tr = product.description_tr,
            # rop.description_ir = product.description_ir,
            # rop.unit = product.unit,
            # rop.weight = product.weight,
            # rop.unit_secondary = product.unit_secondary,
            # rop.price = product.price,
            # #rop.color_making_room_1400 = None,
            rop.avarage_previous_year = average_sale_amount_previous_year
            
            # amount sales from month_1 to month_12
            for month_number in range(1, 13):
                # Filter the `ProductPerformance` objects by the current year and month.
                product_performance_current_month = product_performance.filter(year=jalali_date_now_year, month=month_number)

                # Calculate the total sale amount for the current month.
                total_sale_current_month = product_performance_current_month.aggregate(total_sale=Coalesce(Sum('sale_amount', output_field=models.FloatField()), float(0)))['total_sale']


                # Set the total sale amount for the current month for the corresponding `ROP` month field.
                setattr(rop, f'month_{month_number}', total_sale_current_month)
            rop.total_sale = product_performance.filter(year=jalali_date_now_year).aggregate(total_sale=Coalesce(Sum('sale_amount'), float(0)))['total_sale']
            rop.warehouse = warehouse.stock
            rop.goods_on_the_road = float(0) #! goods on road values will be updated!!!!!!!!!!!!!!!!!!!!!!!!!!!!
            rop.total_stock_all = rop.warehouse+ rop.goods_on_the_road
            rop.monthly_mean = rop.total_sale/int(jalali_date_now.month)
            try: 
                rop.total_month_stock = rop.total_stock_all/rop.monthly_mean
            except Exception as e:
                rop.total_month_stock = 0
            
            # Get the values of `month_1` to the current month.
            values = [getattr(rop, f'month_{i}', 0) for i in range(1, jalali_date_now_month + 1)]
            # Calculate the standard deviation.
            if len(values) > 0:
                standard_deviation = statistics.stdev(values)
            else:
                standard_deviation = 0

            # Set the `standart_deviation` field of the `ROP` instance.
            rop.standart_deviation = standard_deviation
            
            rop.lead_time = 2 #! will be given by user
            rop.product_coverage_percentage = 95 #! will be given by user
            rop.demand_status =  rop.standart_deviation * (rop.lead_time)**0.5

            #safety stock
            try:
                result = norm.ppf(rop.product_coverage_percentage) * rop.demand_status
                result_rounded_up = math.ceil(result)           
                rop.safety_stock = result_rounded_up
            except:
                rop.safety_stock = float(0)

            rop.rop = (rop.lead_time * rop.monthly_mean) + rop.safety_stock 
            #rop.monthly_mean = rop.total_sale/int(jalali_date_now.month)
            rop.new_party = rop.lead_time * rop.monthly_mean
            
            # cycle_service_level
            try:
                # Calculate the PDF of the normal distribution.
                pdf_value = norm.pdf(rop.rop, loc=rop.monthly_mean, scale=rop.demand_status)

                # Set the result to the PDF.
                rop.cycle_service_level = pdf_value
            except:
                # If an error occurs, set the result to 0.
                rop.cycle_service_level = 0
            rop.total_stock = rop.total_stock_all
            
            # need_products
            if rop.rop >= rop.total_stock:
                rop.need_prodcuts = rop.rop + rop.new_party - rop.total_stock
            else:
                 rop.need_prodcuts = 0

            # over stock
            if rop.total_stock_all > (1.2*(rop.safety_stock + rop.new_party)):  #! Stock Over Factor will be declared and it will produced by user
                rop.over_stock = 1
            else:
                rop.over_stock = 0
            
            rop.calculated_need = rop.need_prodcuts #! previous year is needed??????
            
            # calculated max stock #! previous year is needed??????
            try:
                rop.calculated_max_stock = (rop.rop + rop.new_party)/rop.monthly_mean
            except:
                rop.calculated_max_stock = 0

            # calculated min stock #! previous year is needed??????
            try:
                rop.calculated_min_stock = rop.rop /rop.monthly_mean
            except:
                rop.calculated_min_stock = 0
            rop.save()
        except Products.DoesNotExist:
            pass
        except Exception as e:
            return JsonResponse({'error': str(e)})

class ROPView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        print(data)
        product_code = data.get('product_code')
        lead_time = int(data.get('lead_time'))
        service_level = float(data.get('service_level'))
        forecast_period = int(data.get('forecast_period'))
        product_values = ProductPerformance.objects.filter(product_code=product_code)
       
        try:
            last_sales = MonthlyProductSales.objects.filter(product_code=product_code).values("date").latest("date")
            last_sale_date = last_sales["date"].strftime('%Y-%m-%d')
            jalali_date = current_jalali_date()
            jalali_date_str = jalali_date.strftime('%Y-%m-%d').split("-")
        except MonthlyProductSales.DoesNotExist:
            return JsonResponse({"error" : f"There is no product sales data with product code: {product_code} "}, status=400)
        try: 
            warehouse = Warehouse.objects.get(product_code = product_code)
            stock = warehouse.stock
        except Warehouse.DoesNotExist:
            print("sadasdaasd")
            return JsonResponse({"error" : f"There is no product in warehouse with product code: {product_code} "}, status=400)
        
        dates_for_sales = [jdatetime.date(item.year, item.month, 1).strftime('%Y-%m-%d') for item in product_values]
        sales = [item.sale_amount for item in product_values]
        product_values = [[1, item.month, item.year, item.product_code, item.sale_amount] for item in product_values]
        
        holt_all_sales, holt_prev_sales, holt_future_sales, holt_future_stocks, holt_order_flag, holt_safety_stock, holt_rop, holt_order = get_model("holt", True, jalali_date_str, product_code, product_values, stock, lead_time, service_level, 12, forecast_period )
        exp_all_sales, exp_prev_sales, exp_future_sales, exp_future_stocks, exp_order_flag, exp_safety_stock, exp_rop, exp_order = get_model("exp", True, jalali_date_str, product_code, product_values, stock, lead_time, service_level, 12, forecast_period )
        avrg_all_sales, avrg_prev_sales, avrg_future_sales, avrg_future_stocks, avrg_order_flag, avrg_safety_stock, avrg_rop, avrg_order = get_model("average", True, jalali_date_str, product_code, product_values, stock, lead_time, service_level, 12, forecast_period )
        if avrg_order_flag == False:
            avrg_order = 0
        if exp_order_flag == False:
            exp_order = 0
        if holt_order_flag == False:
            holt_order = 0
        avrg_future_forecast_dates = generate_future_forecast_dates(len(avrg_future_sales))
        exp_future_forecast_dates = generate_future_forecast_dates(len(exp_future_sales))
        holt_future_forecast_dates = generate_future_forecast_dates(len(holt_future_sales))
        print("avrg_future_forecast_dates: ", avrg_future_forecast_dates)
        print("holt_future_forecast_dates: ", holt_future_forecast_dates)
        print("holt_order: ", holt_order)
        item = ROP.objects.get(product_code_ir = product_code)
        rop_list = rop_list = [
                item.group,
                item.subgroup,
                item.feature,
                item.new_or_old_product,
                item.related,
                item.origin,
                item.product_code_ir,
                item.product_code_tr,
                item.dont_order_again,
                item.description_tr,
                item.description_ir,
                item.unit,
                item.weight,
                item.unit_secondary,
                item.price,
                item.avarage_previous_year,
                item.month_1,
                item.month_2,
                item.month_3,
                item.month_4,
                item.month_5,
                item.month_6,
                item.month_7,
                item.month_8,
                item.month_9,
                item.month_10,
                item.month_11,
                item.month_12,
                item.total_sale,
                item.warehouse,
                item.goods_on_the_road,
                item.total_stock_all,
                item.total_month_stock,
                item.standart_deviation,
                item.lead_time,
                item.product_coverage_percentage,
                item.demand_status,
                item.safety_stock,
                item.rop,
                item.monthly_mean,
                item.new_party,
                item.cycle_service_level,
                item.total_stock,
                item.need_prodcuts,
                item.over_stock,
                item.calculated_need,
                item.calculated_max_stock,
                item.calculated_min_stock,
            ]
        return JsonResponse({'rop_list': rop_list, 'stock': stock,
                            'avrg_dates_for_sales': dates_for_sales,
                            'avrg_sales': sales,
                            'avrg_future_forecast_dates': avrg_future_forecast_dates, 
                            'avrg_future_sales': avrg_future_sales, 
                            'avrg_order_flag': avrg_order_flag, 
                            'avrg_order': avrg_order,
                            'holt_dates_for_sales': dates_for_sales,
                            'holt_sales': sales,
                            'holt_future_forecast_dates': holt_future_forecast_dates, 
                            'holt_future_sales': holt_future_sales.tolist(), 
                            'holt_order_flag': holt_order_flag, 
                            'holt_order': holt_order,
                            'exp_dates_for_sales': dates_for_sales,
                            'exp_sales': sales,
                            'exp_future_forecast_dates': exp_future_forecast_dates, 
                            'exp_future_sales': exp_future_sales.tolist(), 
                            'exp_order_flag': exp_order_flag, 
                            'exp_order': exp_order }, safe=False)



# endregion

# region OrderList

@receiver(post_save, sender=Sales)
def create_sales_signal(sender, instance, created, **kwargs):
    if created:
        models = ['average', 'holt', 'exp']
        order_flags = {'average': False, 'holt': False, 'exp': False}
        orders = {'average': 0, 'holt': 0, 'exp': 0}
        product_code = instance.product_code
        product_values = ProductPerformance.objects.filter(product_code=product_code)
        product_values = [[1, item.month, item.year, item.product_code, item.sale_amount] for item in product_values]
        weight = Products.objects.get(product_code_ir = product_code).weight
        try:
            last_sales = MonthlyProductSales.objects.filter(product_code=product_code).values("date").latest("date")
            last_sale_date = last_sales["date"].strftime('%Y-%m-%d')
            jalali_date = current_jalali_date()
            jalali_date_str = jalali_date.strftime('%Y-%m-%d').split("-")
        except MonthlyProductSales.DoesNotExist:
            return JsonResponse({"error" : f"There is no product sales data with product code: {product_code} "})
        try: 
            warehouse = Warehouse.objects.get(product_code = product_code)
            stock = warehouse.stock
        except Warehouse.DoesNotExist:
            return JsonResponse({"error" : f"There is no product in warehouse with product code: {product_code} "})

        for model in models:
            all_sales, prev_sales, future_sales, future_stocks, order_flag, safety_stock, rop, order = get_model(
                model, True, jalali_date_str, product_code, product_values, stock, 3, 0.99,
                12, 3
            )

            if order_flag:
                is_active = True
                order_flags[model] = True
                orders[model] = order

        is_active = any(order_flags.values())

        try:
            existing_order_list = OrderList.objects.get(product_code=product_code, is_active=True)
        except OrderList.DoesNotExist:
            existing_order_list = None

        if is_active:
            if existing_order_list:
                # Update the existing OrderList object
                existing_order_list.order_flag_avrg = order_flags['average']
                existing_order_list.order_flag_exp = order_flags['exp']
                existing_order_list.order_flag_holt = order_flags['holt']
                existing_order_list.order_avrg = orders['average']
                existing_order_list.order_exp = orders['exp']
                existing_order_list.order_holt = orders['holt']
                existing_order_list.current_date = jalali_date
                existing_order_list.current_stock = stock
                existing_order_list.weight = weight
                existing_order_list.average_sale = np.mean(all_sales)
                existing_order_list.is_ordered = False
                existing_order_list.save()
            else:
                # Create a new OrderList object
                order_list = OrderList(
                    product_code=product_code,
                    order_flag_avrg=order_flags['average'],
                    order_flag_exp=order_flags['exp'],
                    order_flag_holt=order_flags['holt'],
                    order_avrg=orders['average'],
                    order_exp=orders['exp'],
                    order_holt=orders['holt'],
                    current_date=jalali_date,
                    current_stock=stock,
                    weight=weight,
                    average_sale=np.mean(all_sales),
                    is_active=is_active,
                    is_ordered=False
                )
                order_list.save()




class OrderListView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        order_lists = OrderList.objects.filter(is_active = True).values() #! False True olacak
        order_list_data = [
            [o['id'], o['current_date'].strftime('%Y-%m-%d'), o['product_code'], o['weight'], o['average_sale'], o['current_stock'], o['order_avrg'], o['order_exp'], o['order_holt'],
              o['decided_order']] for o in order_lists
        ]
        return JsonResponse(order_list_data, safe=False)

class EditOrderListView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)
    def post(self, request, *args, **kwargs):
        # Expecting the updated data to be sent as a list of dictionaries
        updated_order_list_data = json.loads(request.body)
        

        try:
            
            order = OrderList.objects.get(id=updated_order_list_data['id'])           
            order.current_date = updated_order_list_data['current_date']
            order.product_code = updated_order_list_data['product_code']
            order.weight = updated_order_list_data['weight']
            order.average_sale = updated_order_list_data['average_sale']
            order.current_stock = updated_order_list_data['current_stock']
            order.order_avrg = updated_order_list_data['order_avrg']
            order.order_exp = updated_order_list_data['order_exp']
            order.order_holt = updated_order_list_data['order_holt']
            order.decided_order = float(updated_order_list_data['decided_order'])

            if order.decided_order > 0:
                order.is_active = False
                order.is_ordered = True

            order.save()
        except OrderList.DoesNotExist:
            return JsonResponse({"error": "Order not found"}, status=404)

        return JsonResponse({'message': "Your order successfully send to the Goods On Road."}, status=200)


# endregion

# region GoodsOnRoad

@receiver(post_save, sender=OrderList)
def update_goods_on_road(sender, instance, created, **kwargs):
    if not created and instance.is_ordered:
        product = Products.objects.get(product_code_ir=instance.product_code)
        goods_on_road, created = GoodsOnRoad.objects.get_or_create(product_code=instance.product_code, is_terminated=False)
        goods_on_road.product_name_tr = product.description_tr
        goods_on_road.product_name_ir = product.description_ir
        goods_on_road.decided_order = instance.decided_order
        goods_on_road.weight = instance.weight
        goods_on_road.truck_id = None
        goods_on_road.is_ordered = instance.is_ordered
        goods_on_road.is_on_truck = False
        goods_on_road.save()

class GoodsOnRoadView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        goods_on_road = GoodsOnRoad.objects.filter(is_terminated=False).values()
        goods_on_road_data = [
            [g['product_code'], g['product_name_tr'], g['product_name_ir'], g['decided_order'], g['weight'], g['truck_name']]
            for g in goods_on_road
        ]
        return JsonResponse(goods_on_road_data, safe=False)

class ApproveProductsToOrderView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            if not data.get("truck_name"):
                return JsonResponse({'error': "Truck Name cannot be empty."}, status=400)

            truck = Trucks.objects.filter(truck_name=data.get("truck_name"), is_ordered=False).first()
            if not truck:
                return JsonResponse({'error': "No active truck found with the given Name."}, status=404)

            if float(data.get("decided_order", 0)) <= 0:
                return JsonResponse({'error': "Decided order cannot be equal to or smaller than zero."}, status=400)

            goods_on_road = GoodsOnRoad.objects.get(product_code=data['product_code'], is_terminated=False)
            goods_on_road.truck_name = data['truck_name']
            goods_on_road.decided_order = data['decided_order']
            goods_on_road.is_terminated = True
            goods_on_road.is_on_truck = True
            goods_on_road.save()

            return JsonResponse({'message': "GoodsOnRoad object updated successfully."}, status=200)
        except GoodsOnRoad.DoesNotExist:
            return JsonResponse({'error': "GoodsOnRoad object not found."}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

# endregion

# region Trucks

class AddTruckView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            # Check if a truck with the same name and is_arrived=False already exists
            existing_truck = Trucks.objects.filter(truck_name=data.get("truck_name"), is_arrived=False)
            if existing_truck.exists():
                return JsonResponse({'error': "A truck with the same name already exists and has not arrived yet."}, status=400)

            estimated_order_date_jalali = data.get("estimated_order_date").split("-")
            estimated_arrival_date_jalali = data.get("estimated_arrival_date").split("-")

            try:
                estimated_order_date = jdatetime.date(int(estimated_order_date_jalali[0]), int(estimated_order_date_jalali[1]), int(estimated_order_date_jalali[2]))
                estimated_arrival_date = jdatetime.date(int(estimated_arrival_date_jalali[0]), int(estimated_arrival_date_jalali[1]), int(estimated_arrival_date_jalali[2]))
            except Exception as e:
                return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD'"}, status=400)

            truck = Trucks(
                truck_name=data.get("truck_name"),
                estimated_order_date=estimated_order_date,
                estimated_arrival_date=estimated_arrival_date,
                is_arrived=False,
                is_ordered= False,
                is_waiting = True
            )
            truck.save()
            return JsonResponse({'message': "Truck added successfully"}, status=200)
        except IndexError:
            return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD' "}, status=400)
        except ValueError:
            return JsonResponse({'error': "The date you entered is in the wrong format. The correct date format is 'YYYY-MM-DD' "}, status=400)
        except Exception as e:
             return JsonResponse({'error': str(e)}, status=500)

class WaitingTrucksView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        waiting_trucks = Trucks.objects.filter(is_waiting=True).values_list('truck_name', flat=True)
        goods_on_road = GoodsOnRoad.objects.filter(is_on_truck=True).values()
        grouped_goods = {}


        for good in goods_on_road:

            if good['truck_name'] not in grouped_goods:

                grouped_goods[good['truck_name']] = []


            grouped_goods[good['truck_name']].append(good)

        return JsonResponse(grouped_goods, safe=False)

class ApproveWaitingTruckView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            truck_name = data.get("truck_name")

            if not truck_name:
                return JsonResponse({'error': "Truck name cannot be empty."}, status=400)

            truck = Trucks.objects.get(truck_name=truck_name, is_waiting=True)

            if not truck:
                return JsonResponse({'error': "No truck found with the given name on waiting list"}, status=404)

            truck.is_waiting = False
            truck.is_ordered = True
            truck.save()
            
            goods_on_road = GoodsOnRoad.objects.filter(truck_name=truck_name, is_on_truck=True)

            for good in goods_on_road:
                good.is_on_truck = False
                good.is_on_road = True
                good.save()

            return JsonResponse({'message': f"Truck with name: {truck_name} is successfully ordered."}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class TrucksOnRoadView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        trucks_on_road = Trucks.objects.filter(is_ordered=True).values_list('truck_name', flat=True)
        goods_on_road = GoodsOnRoad.objects.filter(is_on_road = True).values()
        grouped_goods = {}

        for good in goods_on_road:
            if good['truck_name'] not in grouped_goods:
                grouped_goods[good['truck_name']] = []

            grouped_goods[good['truck_name']].append(good)

        return JsonResponse(grouped_goods, safe=False)
    
class EditGoodsOnRoadView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)

            product_code = data.get('product_code')
            good_on_road = GoodsOnRoad.objects.get(product_code=product_code)

            # Update decided_order
            new_decided_order = float(data.get('new_decided_order'))
            if new_decided_order is not None and new_decided_order >= 0:
                good_on_road.decided_order = new_decided_order
            else:
                return JsonResponse({'error': "Decided Order cannot be negative or empty."}, status=400)

            good_on_road.save()
            return JsonResponse({'message': "Your changes have been successfully saved."}, status=200)

        except GoodsOnRoad.DoesNotExist:
            return JsonResponse({'error': "Good not found."}, status=400)

        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    

class ApproveArrivedTruckView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            truck_name = data.get("truck_name")

            if not truck_name:
                return JsonResponse({'error': "Truck name cannot be empty."}, status=400)

            truck = Trucks.objects.filter(truck_name=truck_name, is_ordered=True).first()
            if not truck:
                return JsonResponse({'error': "No truck found with the given name and is_ordered=True."}, status=404)

            goods_on_road = GoodsOnRoad.objects.filter(truck_name=truck_name, is_on_road=True)
            for good in goods_on_road:
                try:
                    warehouse_product = Warehouse.objects.get(product_code=good.product_code)
                except Warehouse.DoesNotExist:
                   return JsonResponse({'error': f"Product with code {good.product_code} does not exist in the warehouse."}, status=500)

            truck.is_ordered = False
            truck.is_arrived = True
            truck.save()

            goods_on_road = GoodsOnRoad.objects.filter(truck_name=truck_name, is_on_road=True)
            for good in goods_on_road:
                good.is_on_road = False
                good.is_arrived = True
                good.save()

                warehouse_product = Warehouse.objects.get(product_code=good.product_code)
                if warehouse_product:
                    warehouse_product.stock += good.decided_order
                    warehouse_product.save()

            return JsonResponse({'message': "Truck, related GoodsOnRoad objects, and Warehouse stock updated successfully."}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

# endregion


# region Notifications

@receiver(post_save, sender=OrderList)
def create_notification_order_list(sender, instance, created, **kwargs):
    notification_order_list = NotificationsOrderList(
        current_date=instance.current_date,
        product_code=instance.product_code,
        is_active=instance.is_active,
        order_avrg=instance.order_avrg,
        order_exp=instance.order_exp,
        order_holt=instance.order_holt,
    )
    notification_order_list.save()

class NotificationsView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def get(self, request, *args, **kwargs):
        notifications = NotificationsOrderList.objects.filter(is_active=True).values()
        notification_data = [
            [n['id'], n['current_date'].strftime('%Y-%m-%d'), n['product_code'], n['order_avrg'], n['order_exp'], n['order_holt']]
            for n in notifications
        ]
        return JsonResponse(notification_data, safe=False)

class DeleteNotificationView(APIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (JWTAuthentication,)

    def post(self, request, *args, **kwargs):
        notification_id = request.data.get('id')
        if not notification_id:
            return JsonResponse({"error": "Missing 'id' parameter"}, status=400)

        try:
            notification = NotificationsOrderList.objects.get(id=notification_id)
        except NotificationsOrderList.DoesNotExist:
            return JsonResponse({"error": f"No notification found with id: {notification_id}"}, status=404)

        notification.is_active = False
        notification.save()

        return JsonResponse({"success": f"Notification with id {notification_id} has been deactivated"})



# endregion








